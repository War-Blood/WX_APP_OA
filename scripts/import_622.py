import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:/Users/WarBlood/Desktop/6-22数据.xlsx')
ws = wb.active

def esc(v):
    if v is None: return "''"
    s = str(v).replace('\\', '\\\\').replace("'", "\\'")
    return "'" + s + "'"

def dt(v):
    if v is None: return 'NULL'
    if isinstance(v, datetime): return "'" + v.strftime('%Y-%m-%d') + "'"
    s = str(v).strip()
    if s: return "'" + s + "'"
    return 'NULL'

def num(v):
    if v is None: return '0'
    try: return str(int(float(str(v))))
    except: return '0'

def dec(v):
    if v is None: return '0'
    try: return str(float(str(v)))
    except: return '0'

lines = []
count = 0
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if not row or row[0] is None: 
        continue
    if not isinstance(row[0], datetime):
        continue

    rpt      = dt(row[0])
    entry    = dt(row[2]) if 2 < len(row) else 'NULL'
    init_trip= dt(row[3]) if 3 < len(row) else 'NULL'
    project  = esc(row[4]) if 4 < len(row) and row[4] else "''"
    area     = esc(row[5]) if 5 < len(row) and row[5] else "''"
    related  = esc(row[6]) if 6 < len(row) and row[6] else "''"
    w1_raw   = str(row[7]).strip() if 7 < len(row) and row[7] else ''
    w2_raw   = str(row[8]).strip() if 8 < len(row) and row[8] else ''
    if w1_raw and w2_raw:
        workers = esc(w1_raw + '\u3001' + w2_raw)
    elif w1_raw:
        workers = esc(w1_raw)
    elif w2_raw:
        workers = esc(w2_raw)
    else:
        workers = "''"
    machine  = esc(row[9]) if 9 < len(row) and row[9] else "''"
    wc       = num(row[10]) if 10 < len(row) else '0'
    wcontent = esc(row[11]) if 11 < len(row) and row[11] else "''"
    req_qty  = num(row[12]) if 12 < len(row) else '0'
    comp_qty = num(row[13]) if 13 < len(row) else '0'
    progress = dec(row[14]) if 14 < len(row) else '0'
    today    = esc(row[15]) if 15 < len(row) and row[15] else "''"
    tomorrow = esc(row[16]) if 16 < len(row) and row[16] else "''"
    ttype    = esc(row[17]) if 17 < len(row) and row[17] else "''"
    tmtype   = esc(row[18]) if 18 < len(row) and row[18] else "''"
    remark   = esc(row[19]) if 19 < len(row) and row[19] else "''"
    biz_days = num(row[20]) if 20 < len(row) else '0'
    pers_days= num(row[21]) if 21 < len(row) else '0'

    sql = (
        "INSERT INTO daily_reports "
        "(report_date, entry_date, initial_biz_trip_date, project, area, "
        "related_party, workers, machine_model, worker_count, work_content, "
        "required_qty, completed_qty, progress_percent, today_work, tomorrow_plan, "
        "today_work_type, tomorrow_work_type, remark, biz_trip_days, "
        "personal_biz_trip_days, user_id, status) VALUES "
        f"({rpt}, {entry}, {init_trip}, {project}, {area}, "
        f"{related}, {workers}, {machine}, {wc}, {wcontent}, "
        f"{req_qty}, {comp_qty}, {progress}, {today}, {tomorrow}, "
        f"{ttype}, {tmtype}, {remark}, {biz_days}, {pers_days}, 1, 'approved')"
    )
    lines.append(sql)
    count += 1

OUT = r'Y:/AI/WX-APP-OA/sql/import_latest.sql'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('USE wx_app_oa;\n')
    for l in lines:
        f.write(l + ';\n')

print(f'Generated {count} rows -> {OUT}')
